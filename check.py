import openpyxl

import openpyxl
import json

# The official 34 mandals list
mandals_34 = [
    "Balayapalle",
    "Buchinaidu Kandriga",
    "Chandragiri",
    "Chillakur",
    "Chinnagottigallu",
    "Chittamur",
    "Dakkili",
    "Doravarisatram",
    "Gudur",
    "Kota",
    "K. V. B. Puram",
    "Nagalapuram",
    "Naidupet",
    "Narayanavanam",
    "Ojili",
    "Pakala",
    "Pellakur",
    "Pichatur",
    "Puttur",
    "Ramachandrapuram",
    "Renigunta",
    "Satyavedu",
    "Srikalahasti",
    "Sullurpeta",
    "Tada",
    "Thottambedu",
    "Tirupati Rural",
    "Tirupati Urban",
    "Vadamalapeta",
    "Vakadu",
    "Varadaiahpalem",
    "Venkatagiri",
    "Yerpedu",
    "Yerravaripalem"
]

def sanitize(s):
    res = s.lower()
    res = "".join(c for c in res if c.isalnum())
    res = res.replace("thi", "ti")
    res = res.replace("palli", "palle")
    if res.endswith("pet"):
        res = res[:-3] + "peta"
    return res

sanitized_34_map = {sanitize(m): m for m in mandals_34}
# Custom aliases for matching Excel names
aliases = {
    sanitize("Kumara Venkata Bhupala Puram"): "K. V. B. Puram",
    sanitize("Buchi Naidu Kandriga"): "Buchinaidu Kandriga",
    "srikalahasthi": "Srikalahasti",
    "vadamalapet": "Vadamalapeta",
    "varadalahpalem": "Varadaiahpalem",
    "sullurupeta": "Sullurpeta",
    "ozili": "Ojili"
}

path = 'datasets/mandal wise medical facilities.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb['Sheet2']

extracted_data = {}
for r in range(2, 100):
    m_name = sheet.cell(r, 1).value
    if not m_name:
        continue
    m_name_str = str(m_name).strip()
    m_san = sanitize(m_name_str)
    
    # Resolve aliases
    mapped_name = aliases.get(m_san)
    if not mapped_name:
        mapped_name = sanitized_34_map.get(m_san)
        
    if mapped_name:
        vhc = int(sheet.cell(r, 3).value or 0)
        phc = int(sheet.cell(r, 4).value or 0)
        uphc = int(sheet.cell(r, 5).value or 0)
        ahs = int(sheet.cell(r, 6).value or 0)
        chcs = int(sheet.cell(r, 7).value or 0)
        ths = int(sheet.cell(r, 8).value or 0)
        total_facilities = int(sheet.cell(r, 9).value or 0)
        
        # Calculate dedicated hospitals count (AHs + CHCs + THs)
        hospitals_count = ahs + chcs + ths
        
        extracted_data[mapped_name] = {
            "mandal": mapped_name,
            "village_health_clinics_vhc": vhc,
            "primary_health_centers_phc": phc,
            "urban_primary_health_centers_uphc": uphc,
            "area_hospitals_ahs": ahs,
            "community_health_centers_chcs": chcs,
            "teaching_hospitals_ths": ths,
            "hospitals_count": hospitals_count,
            "total_health_facilities": total_facilities
        }

# For any mandals in the 34 list not found in the excel sheet, initialize them with 0
for m in mandals_34:
    if m not in extracted_data:
        extracted_data[m] = {
            "mandal": m,
            "village_health_clinics_vhc": 0,
            "primary_health_centers_phc": 0,
            "urban_primary_health_centers_uphc": 0,
            "area_hospitals_ahs": 0,
            "community_health_centers_chcs": 0,
            "teaching_hospitals_ths": 0,
            "hospitals_count": 0,
            "total_health_facilities": 0
        }

# Convert to list ordered exactly as the 34 mandals list
final_list = [extracted_data[m] for m in mandals_34]

# Save output JSON
output_path = 'datasets/mandal wise no of hospitals.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(final_list, f, indent=2)

print(f"Successfully processed {len(final_list)} mandals!")
print(f"Written output to: {output_path}")















