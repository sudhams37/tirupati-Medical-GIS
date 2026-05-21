import openpyxl
import json
import re

path = 'datasets/Ward-wise-population-hospital-data.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb.active

wards_data = {}
for r in range(2, 200):
    val = sheet.cell(r, 1).value
    if val is None:
        continue
    
    ward_name = sheet.cell(r, 3).value
    population = sheet.cell(r, 4).value
    hospitals = sheet.cell(r, 5).value
    
    if ward_name:
        # Extract number from "Ward No XX" or similar
        match = re.search(r'\d+', str(ward_name))
        if match:
            code = int(match.group())
        else:
            code = r - 1 # Fallback
            
        wards_data[str(code)] = {
            "code": code,
            "name": str(ward_name).strip(),
            "population": int(population or 0),
            "hospitals": int(hospitals or 0)
        }

# Verify and save
output_path = 'datasets/ward_population_hospitals.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(wards_data, f, indent=2)

print(f"Processed {len(wards_data)} wards.")
print(f"Sample data for Code '1': {wards_data.get('1')}")
print(f"Written output to: {output_path}")
