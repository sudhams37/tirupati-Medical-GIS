import openpyxl

path = 'datasets/Ward-wise-population-hospital-data.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb.active

data = []
for r in range(2, 200):
    val = sheet.cell(r, 1).value
    if val is None:
        continue
    city = sheet.cell(r, 2).value
    ward = sheet.cell(r, 3).value
    pop = sheet.cell(r, 4).value
    hosp = sheet.cell(r, 5).value
    
    if pop is not None and hosp is not None:
        pop = int(pop)
        hosp = int(hosp)
        ratio = pop / hosp if hosp > 0 else float('inf')
        data.append({
            'ward': ward,
            'population': pop,
            'hospitals': hosp,
            'ratio': ratio
        })

print(f"Total wards loaded: {len(data)}")
ratios = [d['ratio'] for d in data if d['ratio'] != float('inf')]
ratios.sort()
print(f"Min ratio (pop/hosp): {min(ratios) if ratios else 'N/A'}")
print(f"Max ratio (pop/hosp): {max(ratios) if ratios else 'N/A'}")
print(f"Median ratio (pop/hosp): {ratios[len(ratios)//2] if ratios else 'N/A'}")
print(f"Average ratio (pop/hosp): {sum(ratios)/len(ratios) if ratios else 'N/A'}")

# Print first 20 wards
for d in data[:20]:
    print(f"{d['ward']}: pop={d['population']}, hosp={d['hospitals']}, ratio={d['ratio']:.1f}")
