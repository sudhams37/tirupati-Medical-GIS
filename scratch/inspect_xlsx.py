import openpyxl

path = 'datasets/Ward-wise-population-hospital-data.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
print("Sheet names:", wb.sheetnames)

sheet = wb.active
print("Active sheet title:", sheet.title)

for r in range(1, 15):
    row_vals = [sheet.cell(r, c).value for c in range(1, 15)]
    # Filter out trailing Nones for cleaner display
    while row_vals and row_vals[-1] is None:
        row_vals.pop()
    if row_vals:
        print(f"Row {r}: {row_vals}")
