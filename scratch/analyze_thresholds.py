import openpyxl

path = 'datasets/Ward-wise-population-hospital-data.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb.active

data = []
for r in range(2, 200):
    val = sheet.cell(r, 1).value
    if val is None:
        continue
    ward = sheet.cell(r, 3).value
    pop = int(sheet.cell(r, 4).value or 0)
    hosp = int(sheet.cell(r, 5).value or 0)
    ratio = pop / hosp if hosp > 0 else float('inf')
    data.append({'ward': ward, 'pop': pop, 'hosp': hosp, 'ratio': ratio})

thresholds = [5000, 8000, 10000, 12000, 15000, 20000]
for t in thresholds:
    sufficient = sum(1 for d in data if d['ratio'] <= t)
    insufficient = len(data) - sufficient
    print(f"Threshold: 1 hosp per {t} pop => Sufficient: {sufficient}, Insufficient: {insufficient}")
